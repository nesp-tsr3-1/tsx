from pyproj import Transformer
from tsx.db import T1Survey, T1Sighting, T1Site, T2Survey, T2Sighting, T2Site, Taxon, TaxonLevel, Source, DataImport, SourceType, SearchType, Unit, UnitType, Management, ProjectionName, DataProcessingType, get_session, MonitoringProgram
import tsx.util
import os
import logging
import sys
from datetime import date, datetime, time
import argparse
from contextlib import contextmanager
import csv
from sqlalchemy.orm.exc import MultipleResultsFound, NoResultFound
from sqlalchemy.exc import IntegrityError
from shapely.geometry import Point
from tqdm import tqdm
from sqlalchemy import func
import openpyxl
import zipfile
import re
from functools import lru_cache

from six import text_type

# Ignore MySQL warning caused by binary geometry data
import warnings
warnings.filterwarnings("ignore", ".*Invalid utf8 character string.*")

log = logging.getLogger(__name__)

# Helper to make logging and progress bar work together
class TqdmStream(object):
	def write(self, x):
		tqdm.write(x.strip())
	def flush(self):
		pass

def main():
	logging.basicConfig(stream=TqdmStream(), level=logging.INFO, format='%(asctime)-15s %(levelname)-8s %(message)s')

	parser = argparse.ArgumentParser(description='Import Type 1/2 survey data into TSX database')
	parser.add_argument('filename', type=str, help='data file to import (Excel/CSV)')
	parser.add_argument('-t', action='store_true', dest='test', help='test database connection')
	parser.add_argument('--type', dest='data_type', choices=[1,2], type=int, help='Type of data (1 or 2)')
	parser.add_argument('-c', action='store_true', dest='commit', help='commit changes (default is dry-run)')
	parser.add_argument('--simple', action='store_true', dest='simple_mode', help='Simple mode')
	args = parser.parse_args()

	if args.filename:
		if not args.simple_mode and not args.data_type:
			parser.error('--type is required')
			return
		if args.simple_mode:
			if args.data_type == 2:
				parser.error('--type 2 is not support in simple mode')
				return
			else:
				args.data_type = 1

		importer = Importer(args.filename, data_type = args.data_type, commit = args.commit, simple_mode = args.simple_mode)
		importer.ingest_data()
	elif args.test:
		test_db()
	else:
		parser.print_help()

def test_db():
	"""
	Test the db conneciton
	"""
	print("Testing DB Connection")
	print()
	session = get_session()
	# list all units
	for u in session.query(Unit).all():
		print("%d: %s" % (u.id, u.description))
	print()
	print("DB Connection successful")


class ImportLogger(logging.LoggerAdapter):
	def process(self, msg, kwargs):
		return 'Row %s: %s' % (self.extra['row_index'], msg), kwargs

class ImportError(Exception):
	pass

class Importer:
	def __init__(self, filename, data_type = None, commit = False, logger = log, progress_callback = None, source_id = None, data_import_id = None, simple_mode = False):
		if data_type not in (1,2):
			raise ValueError("Invalid type (must be 1 or 2)")

		self.messages = []
		self.filename = filename
		self.commit = commit
		self.log = logger
		self.data_type = data_type
		self.progress_callback = progress_callback
		self.simple_mode = simple_mode
		# see check_survey_consistency
		self.survey_fields_by_pk = {}

		# count errors/warnings
		self.log_counter = tsx.util.CounterHandler()
		self.log.addHandler(self.log_counter)
		self.error_count = 0
		self.warning_count = 0
		self.processed_rows = 0

		# used to cache lookups
		self.cache = {}
		self.pending_sightings = []

		# fast mode doesn't update sightings, just inserts
		self.fast_mode = True

		# list of columns that must not be blank if present
		self.non_empty_keys = [
			'SearchTypeDesc',
			'SourceDesc',
			'DataProcessingType',
			'SiteName',
			'TaxonID',
			'Count',
			'StartDate',
			'X',
			'Y',
			'SourcePrimaryKey',
			'UnitID',
			'UnitOfMeasurement',
			'UnitType',
			'ManagementCategory'
		]

		# List of columns that must contain the same value for all rows
		if source_id:
			self.constant_keys = [
				'SourceProvider',
				'SourceDesc',
				'SourceType',
				'SourceDescDetails',
				'MonitoringProgram',
				'MonitoringProgramComments'
			]
		else:
			self.constant_keys = []

		if self.data_type == 2:
			self.non_empty_keys.remove('SiteName')
			self.non_empty_keys.remove('Count')
			self.non_empty_keys.remove('TaxonID')
		if self.simple_mode:
			self.non_empty_keys.remove('TaxonID')

		self.sighting_keys = set(["SpNo", "TaxonID", "Breeding", "Count", "UnitID", "UnitType", "SightingComments"])

		self.source_id = source_id

		self.data_import_id = data_import_id
		self.first_row = None


	def progress_wrapper(self, iterable):
		# Show a progress bar only if we are running as a script
		if __name__ == '__main__':
			return tqdm(iterable, total = self.row_count)
		else:
			return iterable

	def ingest_data(self):
		"""
		This method reads a csv file, and parse it to ingest_row to process each row
		"""

		self.log.info("Starting import")

		session = get_session()
		self.session = session

		try:
			self.clear_previous_data(session)

			# Autodetect excel files
			if zipfile.is_zipfile(self.filename):
				self.log.info('File format: Excel Spreadsheet')
				format = 'excel'
			else:
				self.log.info('File format: CSV')
				format = 'csv'

			if format == 'excel':
				# The processing code generally expects data to come in as strings
				# We could convert everything to str, but that gives us issues with date formats and 'None'
				# So we just convert numeric values to strings
				def normalize_excel_value(v):
					if type(v) == int or type(v) == float:
						return str(v)
					else:
						return v

				wb = openpyxl.load_workbook(open(self.filename, 'rb'), read_only=True)
				for name in wb.sheetnames:
					# Skip any worksheets with 'template' in the name
					# if "template" in name.lower():
						# continue

					ws = wb[name]

					#self.row_count = ws.max_row # max_row cannot be trusted
					self.row_count = len(list(ws.iter_rows()))

					for i, row in self.progress_wrapper(enumerate(ws.rows)):
						row = [normalize_excel_value(cell.value) for cell in row]
						if i == 0:
							headers = row
							self.check_headers(row)
							self.processed_rows = 1
						else:
							# Fill missing cells with None
							if len(row) < len(headers):
								row = row + [None] * (len(headers) - len(row))
							self.process_row(session, dict(zip(headers, row)), i + 1)

					break

			elif format == 'csv':
				# Count rows first - this is used so we can track progress
				with open(self.filename, "r") as csvfile:
					reader = csv.DictReader(csvfile)
					self.row_count = sum(1 for row in reader)

				with open(self.filename, "r") as csvfile:
					reader = csv.DictReader(csvfile)
					self.check_headers(reader.fieldnames)
					self.processed_rows = 1
					for i, row in enumerate(self.progress_wrapper(reader)):
						self.process_row(session, row, i + 2)

			try:
				self.flush_sightings(session) # flush any left over sightings not yet committed
			except Exception as ex:
				self.handle_import_exception(ex)

		except ImportError as e:
			self.log.error(str(e))
			self.log.info("Aborting import - no changes will be committed to the database")
			session.rollback()

		else:
			if self.commit:
				self.log.info("Committing changes")
				session.commit()
			else:
				self.log.info("Not committing changes - dry run only")
				session.rollback()

		self.update_log_counts()

		self.log.info("Processed %s/%s rows" % (self.processed_rows, self.row_count))
		self.log.info("Import processing complete. Errors: %s, Warnings: %s" % (self.error_count, self.warning_count))

	def clear_previous_data(self, session):
		if self.source_id:
			self.log.info('Deleting all previous data for this source')
			session.execute('DELETE FROM t1_survey WHERE source_id = :source_id', { 'source_id': self.source_id })
			session.execute('DELETE FROM t1_site WHERE source_id = :source_id', { 'source_id': self.source_id })
			session.execute('DELETE FROM t2_survey WHERE source_id = :source_id', { 'source_id': self.source_id })
			session.execute('DELETE FROM t2_site WHERE source_id = :source_id', { 'source_id': self.source_id })


	def update_log_counts(self):
		self.error_count = self.log_counter.count('ERROR')
		self.warning_count = self.log_counter.count('WARNING')

	def check_headers(self, headers):
		header_info = [
			"SourcePrimaryKey*",
			"SourceType*",
			"SourceDesc*",
			"SourceDescDetails",
			"SourceProvider",
			"DataProcessingType*",
			"SearchTypeDesc*",
			"LocationName",
			"SiteName*",
			"Y*",
			"X*",
			"ProjectionReference*",
			"PositionalAccuracyInM",
			"StartDate*",
			"FinishDate*",
			"StartTime",
			"FinishTime",
			"DurationInMinutes",
			"DurationInDays/Nights",
			"NumberOfTrapsPerDay/Night",
			"AreaInM2",
			"LengthInKm",
			"SurveyComments",
			"MonitoringProgram",
			"MonitoringProgramComments",
			"ManagementCategory*",
			"ManagementCategoryComments",
			"TaxonID*",
			"CommonName",
			"ScientificName*",
			"Count*",
			"UnitOfMeasurement*",
			"UnitType*",
			"SightingComments"
		]

		required_headers = set(x[0:-1] for x in header_info if x.endswith("*"))
		optional_headers = set(x for x in header_info if not x.endswith("*"))

		# Legacy columns
		optional_headers |= set([
			"SpNo",
			"UnitID",
			"Breeding"
		])

		if self.data_type == 2:
			required_headers -= set(['SourceType', 'TaxonID'])
			required_headers |= set(['SecondarySourceID'])
			optional_headers.add('TaxonID')

		missing_headers = required_headers - set(headers)

		if len(missing_headers) > 0:
			raise ImportError("Missing required columns(s): %s" % ', '.join(missing_headers))

		if set(["UnitID", "UnitOfMeasurement"]).isdisjoint(set(headers)):
			raise ImportError("Either UnitID or UnitOfMeasurement column is required")

		if set(["DurationInMinutes", "DurationInDays/Nights"]).isdisjoint(set(headers)):
			raise ImportError("Either DurationInMinutes or DurationInDays/Nights column is required")

		unrecognized_headers = set(headers) - required_headers - optional_headers - set([None])

		if len(unrecognized_headers) > 0:
			self.log.warning("Unrecognized column(s) - will be ignored: %s" % ', '.join(unrecognized_headers))

		if self.source_id:
			self.log.info("Note: SourceDesc, SourceType, SourceProvider and MonitoringProgram will be ignored since this information is specified via the web interface")


	def process_row(self, session, row, row_index):
		try:
			self.ingest_row(session, row, row_index)
		except Exception as ex:
			self.handle_import_exception(ex)
		finally:
			self.update_log_counts()
			if self.progress_callback:
				self.progress_callback(row_index, self.row_count)

			self.processed_rows += 1


		if self.processed_rows > 100 and self.error_count + self.warning_count > self.processed_rows / 100.0:
			raise ImportError("Warning/error rate too high")

	def handle_import_exception(self, ex):
		if isinstance(ex, ImportError):
			raise ex

		# Detect duplicate source primary key and give a more friendly error message
		if isinstance(ex, IntegrityError):
			msg = ex.orig.args[1] # Took me along time to figure out this line. Doesn't seem to be documented anywhere.
			if 'Duplicate entry' in msg and 'source_primary_key_UNIQUE' in msg:
				raise ImportError("Error: %s [NOTE: Please ensure that your data is sorted by the SourcePrimaryKey column]" % msg)

		self.log.exception("Fatal error during import")
		raise ImportError("Unexpected error - probably a bug (see above)")

	def ingest_row(self, session, row, row_index):
		"""
		This function ingests data one row at a time. Each row is a dictionary

		See check_headers() for a list of field names
		"""

		# Include row number in all log messages
		log = ImportLogger(self.log, { 'row_index': row_index })

		log.debug("Ingesting: %s" % row.values())

		if self.data_type == 1:
			Site, Survey, Sighting = T1Site, T1Survey, T1Sighting
		else:
			Site, Survey, Sighting = T2Site, T2Survey, T2Sighting

		# If this gets set to false, we don't insert the row.
		# We are using a list so that we can update it from a nested function (in Python 3 we would just use 'nonlocal')
		ok = [True]

		# Helper to avoid repetition when checking fields and logging errors
		@contextmanager
		def field(key):
			try:
				yield row.get(key)
			except ValueError as e:
				log.error('%s: [%s] %s' % (key, row.get(key), str(e)))
				ok[0] = False

		# Strip leading/trailing whitespace from all values, and convert empty strings to None
		for key in row:
			if type(row[key]) in (str, text_type):
				row[key] = row[key].strip()
			if row[key] == '':
				row[key] = None

		# Check that survey information doesn't change for the same SourcePrimaryKey
		if self.source_id == None:
			primary_key = row.get('SourcePrimaryKey')
		else:
			if 'SourcePrimaryKey' in row:
				primary_key = "%s__%s" % (self.source_id, row.get('SourcePrimaryKey'))
			else:
				primary_key = "%s__Row_%s" % (self.source_id, row_index)

		try:
			self.check_survey_consistency(row, primary_key)
		except ValueError as e:
			log.error(str(e))
			self.commit = False # serious error
			ok[0] = False

		# Check if there is not sighting data at all (i.e. survey with no sightings)
		sighting_empty = all(row.get(key) == None for key in self.sighting_keys)

		# Check for empty values
		for key in self.non_empty_keys:
			if key in row and row.get(key) == None and not (sighting_empty and key in self.sighting_keys):
				log.error("%s: must not be empty" % key)
				return False

		# Check constant values
		if self.first_row == None:
			self.first_row = row
		else:
			for key in self.constant_keys:
				if row.get(key) != self.first_row.get(key):
					log.error("%s: must match the first row of the file (%s)" % (key, self.first_row,get(key)))
					ok[0] = False

		# Allow N/A for some columns and convert to 'None'
		# Importantly, we do this after checking for empty values
		for key in ['ManagementCategory', 'UnitType', 'DataProcessingType', 'FinishDate']:
			if row.get(key) == 'N/A':
				row[key] = None
		if row.get('ProjectionReference') == 'N/A':
			row['ProjectionReference'] = 'EPSG:4326'

		# Source
		source_type = None
		with field('SourceType') as value:
			if value:
				source_type = validate(value, self.validate_lookup(session, SourceType))

		if row.get('DataProcessingType'):
			data_processing_type = get_or_create(session, DataProcessingType, description=row.get('DataProcessingType'))
		else:
			data_processing_type = None

		if self.source_id:
			# Import via web interface
			source = session.query(Source).get(self.source_id)
			if row_index == 2:
				# source.description = row.get('SourceDesc')
				source.notes = row.get('SourceDescDetails')
				# source.source_type = source_type
				source.data_processing_type = data_processing_type
				# source.provider = row.get('SourceProvider')
				# source.monitoring_program = self.get_or_create_monitoring_program(session, row.get('MonitoringProgram'))
				source.monitoring_program_comments = row.get('MonitoringProgramComments')

				session.flush()

		else:
			# Import via command line
			source = self.get_source(session, row.get('SourceDesc'))

			if source == None:
				source = Source(
					description = row.get('SourceDesc'),
					provider = row.get('SourceProvider'),
					notes = row.get('SourceDescDetails'),
					source_type = source_type,
					data_processing_type = data_processing_type,
					monitoring_program = self.get_or_create_monitoring_program(session, row.get('MonitoringProgram')),
					monitoring_program_comments = row.get('MonitoringProgramComments'))
				session.add(source)
				session.flush()
			else:
				# Existing source
				if source.source_type != source_type:
					log.error("SourceType: doesn't match database for this source")
					ok[0] = False
				if source.provider != row.get('SourceProvider'):
					log.error("SourceProvider: doesn't match database for this source")
					ok[0] = False

		# Data Import
		if self.data_import_id == None:
			data_import = None
		else:
			data_import = session.query(DataImport).get(self.data_import_id)

		# SearchType
		search_type = self.get_or_create_search_type(session, row.get('SearchTypeDesc'))

		with field('ManagementCategory') as value:
			management = validate(value, self.validate_lookup(session, Management))

		site = None
		# Site
		if self.data_type == 1 or row.get('SiteName') != None:
			last_site = self.cache.get('last_site')
			if last_site != None and last_site.name == row.get('SiteName') and last_site.search_type == search_type and last_site.source == source and (self.data_type != 1 or last_site.management == management):
				# Same site as last row - no need to process site
				site = last_site
			else:
				try:
					site_params = {
						'name': row.get('SiteName'),
						'search_type': search_type,
						'source': source,
						'data_import': data_import
					}
					if self.data_type == 1:
						site_params['management'] = management # This property only present for type 1 sites
						site_params['management_comments'] = row.get('ManagementCategoryComments')

					site = get_or_create(session, Site, **site_params)

				except MultipleResultsFound:
					log.error("Found duplicate sites in DB - this must be fixed before import can continue")
					self.commit = False # serious error
					return False

				self.cache['last_site'] = site

		# Survey
		last_survey = self.cache.get('last_survey')
		if last_survey != None and last_survey.source_primary_key == primary_key:
			# Same survey as last row - no need to process survey fields (this yields a huge speed up for type 2/3 data)
			survey = last_survey
		else:
			if self.fast_mode:
				survey = None
			else:
				survey = session.query(Survey).filter_by(source_primary_key = primary_key, data_import = data_import).one_or_none()

			if survey == None:
				survey = Survey(site = site, source_primary_key = primary_key, data_import = data_import)

			self.cache['last_survey'] = survey

			survey.search_type = search_type
			survey.source = source

			if self.data_type == 2:
				survey.secondary_source_id = row.get('SecondarySourceID')

			# Start/Finish date/time
			with field('StartDate') as value:
				survey.start_date_d, survey.start_date_m, survey.start_date_y = parse_date(value, log)

			with field('FinishDate') as value:
				survey.finish_date_d, survey.finish_date_m, survey.finish_date_y = parse_date(value, log)

			with field('StartTime') as value:
				survey.start_time = parse_time(value)

			with field('FinishTime') as value:
				survey.finish_time = parse_time(value)

			if row.get('FinishDate') and (survey.start_date_y, survey.start_date_m, survey.start_date_d) > (survey.finish_date_y, survey.finish_date_m, survey.finish_date_d):
				log.error("Start date after finish date: %s > %s" % ((survey.start_date_y, survey.start_date_m, survey.start_date_d), (survey.finish_date_y, survey.finish_date_m, survey.finish_date_d)));
				ok[0] = False

			# Duration, area, length, location, accuracy

			if row.get('DurationInMinutes') is not None and row.get('DurationInDays/Nights') is not None:
				log.error("Only one of DurationInMinutes or DurationInDays/Nights may be specified, not both")
				ok[0] = False

			with field('DurationInMinutes') as value:
				if value is not None:
					survey.duration_in_minutes = validate(value, validate_int, validate_greater_than(0))

			with field('DurationInDays/Nights') as value:
				if value is not None:
					survey.duration_in_minutes = int(validate(value, validate_float, validate_greater_than(0)) * 24 * 60)

			with field('NumberOfTrapsPerDay/Night') as value:
				if self.data_type == 1 and value is not None:
					survey.number_of_traps = validate(value, validate_int, validate_greater_than(0))

			with field('AreaInM2') as value:
				survey.area_in_m2 = validate(value, validate_float, validate_greater_than(0))

			with field('LengthInKm') as value:
				survey.length_in_km = validate(value, validate_float, validate_greater_than(0))

			with field('PositionalAccuracyInM') as value:
				survey.positional_accuracy_in_m = validate(value, validate_int, validate_greater_than_or_equal(0))

			survey.location = row.get('LocationName')
			survey.comments = row.get('SurveyComments')

			# Coordinates
			with field('X') as value:
				x = validate(value, validate_float)

			with field('Y') as value:
				y = validate(value, validate_float)

			if x != None and y != None:
				if x == 0 or y == 0:
					log.warning('Suspicious zero coordinate before projection: %s, %s' % (x, y))

				projection_ref = self.get_projection_ref(session, row.get('ProjectionReference'))
				coords = create_point(x, y, projection_ref)

				# This produces the necessary SQL to insert WKB geometry with SQLAlchemy.
				# Simpler than using GeoAlchemy in the end, which wasn't compatible with python-mysql-connector anyway.
				survey.coords = func.ST_GeomFromWKB(coords.to_wkb())

				x, y = coords.x, coords.y

				if x < -180 or x > 180 or y < -90 or y > 90:
					log.error('Invalid coordinates after projection: %s, %s' % (x, y))
					ok[0] = False
				elif x == 0 or y == 0:
					log.warning('Suspicious zero coordinate after projection: %s, %s' % (x, y))


		if not sighting_empty:
			# Taxon
			# There are a few different ways to identify the taxon, which are tried in the following order:
			# 1. Taxon ID
			# 2. SpNo (legacy)
			# 3. Scientific name
			# 4. Common name
			taxon_id = row.get('TaxonID')
			spno = None if self.simple_mode else row.get('SpNo')
			common_name = normalize(row.get('CommonName'))
			scientific_name = normalize(row.get('ScientificName'))

			if taxon_id is not None:
				taxon = self.get_taxon(session, taxon_id)
				if not taxon:
					if self.simple_mode:
						taxon = self.create_taxon(session, taxon_id, scientific_name, common_name or 'Unknown')
					else:
						log.error("Invalid TaxonID: %s" % taxon_id)
					return False
			else:
				taxon = None

			if taxon is None and args.simple_mode:
				taxon = self.create_taxon(session, None, scientific_name, common_name or 'Unknown')

			if spno:
				if taxon is None:
					# Get Taxon ID from SpNo
					taxon = self.get_species_taxon(session, spno)
					if taxon is None:
						log.error("Invalid Spno: %s" % spno)
						return False
				else:
					if str(taxon.spno) != spno:
						log.error("Invalid TaxonID/Spno: %s, %s" % (taxon_id, spno))
						return False

			if taxon is None and common_name:
				taxon = self.get_taxon_by_common_name(session, common_name)
				if not taxon:
					log.warning("Unrecognized common name: %s", common_name)

			if taxon is None and scientific_name:
				taxon = self.get_taxon_by_scientific_name(session, scientific_name)
				if not taxon:
					log.warning("Unrecognized scientific name: %s", scientific_name)

			if taxon is None:
				log.error("Could not identify taxon from TaxonID, CommonName or ScientificName")
				return False

			# Now check consistency of common/scientific names
			if common_name and taxon.common_name and taxon.common_name != common_name:
				log.warning("Common name (%s) does not match expected (%s)" % (common_name, taxon.common_name))
			if scientific_name and taxon.scientific_name != scientific_name:
				log.warning("Scientific name (%s) does not match expected (%s)" % (scientific_name, taxon.scientific_name))

			if self.fast_mode:
				sighting = None
			else:
				sighting = session.query(Sighting).filter_by(survey = survey, taxon_id = taxon.id).one_or_none()

			if sighting == None:
				sighting = Sighting(survey = survey, taxon_id = taxon.id)

			with field('Breeding') as value:
				if value in (None, '0'):
					sighting.breeding = False
				elif value == '1':
					sighting.breeding = True
				else:
					raise ValueError('Invalid value (should be 0/blank or 1)')

			with field('Count') as value:
				if value is None:
					sighting.count = 0
				else:
					sighting.count = validate(value, validate_float, validate_greater_than_or_equal(0))

			# Unit
			if (row.get('UnitID') is None) == (row.get('UnitOfMeasurement') is None):
				log.error("Only one of UnitID and UnitOfMeasurement may be specified (and not both)")
				return False

			with field('UnitID') as value:
				if value is not None:
					unit_id = validate(value, validate_int)
					try:
						sighting.unit = self.get_unit(session, unit_id)
					except NoResultFound:
						raise ValueError('Unrecognized ID')

			with field('UnitOfMeasurement') as value:
				if value == 'Sample: Occupancy (# presences/# absences)':
					raise ValueError('Sample: Occupancy (# presences/# absences) is no longer a permitted unit. Instead, please use: Sample: Occupancy (# presences/# surveys)')
				if value is not None:
					sighting.unit = self.get_or_create_unit(session, value)

			with field('UnitType') as value:
				sighting.unit_type = validate(value, self.validate_lookup(session, UnitType))

			sighting.comments = row.get('SightingComments')

		if ok[0]:
			session.add(survey)
			self.pending_sightings.append(sighting)
			# session.add(sighting)

		# Performance optimisation: instead of adding sightings to the session straight away, we periodically flush them in
		# batches, using the ultra-fast 'bulk_save_objects' method
		if row_index % 4000 == 0:
			self.flush_sightings(session)
			# session.flush()


	def flush_sightings(self, session):
		session.flush()
		for sighting in self.pending_sightings:
			sighting.survey_id = sighting.survey.id
			sighting.unit_id = sighting.unit.id
			sighting.unit_type_id = sighting.unit_type.id if sighting.unit_type else None
		session.bulk_save_objects(self.pending_sightings)
		self.pending_sightings = []

	# Helpers that cache results to speed things up:

	def get_species_taxon(self, session, spno):
		if 'species_taxon' not in self.cache:
			self.cache['species_taxon'] = { taxon.spno: taxon for taxon in session.query(Taxon).all() if taxon.taxon_level is not None and taxon.taxon_level.description == 'sp' }
		return self.cache['species_taxon'].get(int(spno))

	def get_taxon(self, session, taxon_id):
		if 'taxa' not in self.cache:
			self.cache['taxa'] = { taxon.id: taxon for taxon in session.query(Taxon).all() }
		return self.cache['taxa'].get(taxon_id)

	def get_taxon_by_common_name(self, session, common_name):
		return self.get_cached('taxon_by_common_name', common_name,
			lambda: session.query(Taxon).filter(Taxon.common_name == common_name).one_or_none())

	def get_taxon_by_scientific_name(self, session, scientific_name):
		return self.get_cached('taxon_by_scientific_name', scientific_name,
			lambda: session.query(Taxon).filter(Taxon.scientific_name == scientific_name).one_or_none())

	def create_taxon(self, session, taxon_id, scientific_name, common_name):
		if taxon_id is None:
			taxon_id = session.execute("""SELECT CONCAT('a', LPAD(COALESCE(SUBSTR((SELECT MAX(id) FROM taxon WHERE id LIKE 'a%'), 2), 0) + 1, 5, '0'));""")

		if scientific_name is None:
			scientific_name = taxon_id

		taxon = Taxon(
					id = taxon_id,
					ultrataxon = True,
					common_name = normalize(common_name),
					scientific_name = normalize(scientific_name),
					taxonomic_group = 'Unknown'
				)
		try:
			del self.cache['taxa']
		except KeyError:
			pass
		session.add(taxon)
		session.flush()



	def validate_lookup(self, session, model):
		lookup = self.get_cached('lookup', model, lambda: { x.description: x for x in session.query(model).all() })
		return validate_lookup(lookup)

	def get_unit(self, session, unit_id):
		return self.get_cached('unit_by_id', unit_id,
			lambda: session.query(Unit).filter(Unit.id==unit_id).one())

	def get_or_create_unit(self, session, description):
		return self.get_cached('unit', description,
			lambda: get_or_create(session, Unit, description = description))

	def get_source(self, session, description):
		return self.get_cached('source', description,
			lambda: session.query(Source).filter_by(description = description).one_or_none())

	def get_source_type(self, session, description):
		if description == None:
			return None
		return self.get_cached('source_type', description,
			lambda: session.query(SourceType).filter_by(description = description).one_or_none())

	def get_or_create_search_type(self, session, description):
		return self.get_cached('search_type', description,
			lambda: get_or_create(session, SearchType, description = description))

	def get_or_create_monitoring_program(self, session, description):
		if description == None:
			return None
		return self.get_cached('monitoring_program', description,
			lambda: get_or_create(session, MonitoringProgram, description = description))

	def get_projection_ref(self, session, projection_name_or_ref):
		return self.get_cached('projection_name', projection_name_or_ref,
			lambda: self.get_projection_ref_uncached(session, projection_name_or_ref))

	def get_projection_ref_uncached(self, session, projection_name_or_ref):
		epsg_srid = session.query(ProjectionName.epsg_srid).filter(ProjectionName.name == projection_name_or_ref).one_or_none()
		return "EPSG:%s" % epsg_srid[0] if epsg_srid else projection_name_or_ref

	def get_cached(self, group, key, fn, cacheNone = False):
		if group not in self.cache:
			self.cache[group] = {}
		g = self.cache[group]

		if key in g:
			return g[key]

		result = fn()

		if result != None or cacheNone:
			g[key] = result

		return result


	# Check that survey details don't change for the same 'SourcePrimaryKey'
	def check_survey_consistency(self, row, primary_key):
		survey_keys = ['SourceType', 'SourceDesc', 'SourceProvider', 'SiteName', 'SearchTypeDesc', 'StartDate', 'FinishDate', 'StartTime', 'FinishTime', 'DurationInMinutes', 'AreaInM2', 'LengthInKm', 'LocationName', 'Y', 'X', 'ProjectionReference', 'PositionalAccuracyInM']
		fields = {key:row.get(key) for key in survey_keys if key in row}

		if primary_key not in self.survey_fields_by_pk:
			self.survey_fields_by_pk[primary_key] = fields
		elif self.survey_fields_by_pk[primary_key] != fields:

			# Find the fields that are different and log them:
			a, b = self.survey_fields_by_pk[primary_key], fields
			a_diff = { k: a[k] for k in a if a[k] != b.get(k) }
			b_diff = { k: b[k] for k in b if b[k] != a.get(k) }

			raise ValueError("Survey fields do not match for the same SourcePrimaryKey (%s):\n%s\n%s" % (row.get('SourcePrimaryKey'), a_diff, b_diff))

@lru_cache(maxsize=None)
def get_proj_transformer(projection_ref):
	return Transformer.from_crs(projection_ref, "EPSG:4326", always_xy=True)

def create_point(x, y, projection_ref):
	if projection_ref not in (None, 'EPSG:4326'):
		try:
			transformer = get_proj_transformer(projection_ref)
		except:
			log.exception("Invalid/unrecognized projection")
			raise ImportError("Invalid/unrecognized projection: %s" % projection_ref)
		x, y = transformer.transform(x, y)

	return Point(x, y)


# see: https://stackoverflow.com/questions/2546207/does-sqlalchemy-have-an-equivalent-of-djangos-get-or-create
def get_or_create(session, model, **kwargs):
	"""
	Gets a single row from the database matching the passed criteria, or creates such a row if none exists
	"""
	instance = session.query(model).filter_by(**kwargs).one_or_none()
	if instance != None:
		return instance
	else:
		instance = model(**kwargs)
		session.add(instance)
		session.flush()
		return instance

def description_lookup(session, model):
	return { x.description: x for x in session.query(model).all() }

# ----- Validation helpers ------
#
# Use like this:
#
# checked_value = validate(raw_value, validate_int, validate_greater_than(10))
#
# Validator raise ValueErrors as appropriate
# None values are passed through without validation, and a validator can also return None to 
#

def validate(value, *validators):
	for validator in validators:
		if value != None:
			value = validator(value)
	return value

def validate_int(x):
	try:
		return int(x)
	except ValueError:
		raise ValueError('must be an integer')

def validate_float(x):
	try:
		return float(x)
	except ValueError:
		raise ValueError('must be a floating point number')

def validate_greater_than(y):
	return validate_condition(lambda x: x > y, 'must be greater than %s' % y)

def validate_greater_than_or_equal(y):
	return validate_condition(lambda x: x >= y, 'must be greater than or equal to %s' % y)

def validate_lookup(lookup):
	def v(x):
		if x in lookup:
			return lookup[x]
		else:
			raise ValueError('must be one of %s' % quoted_strings(lookup.keys()))
	return v

def validate_condition(fn, message):
	def v(x):
		if fn(x):
			return x
		else:
			raise ValueError(message)
	return v

# ----- End Validation Helpers -----

def parse_date(raw_date, log):
	"""
	Parses date in DD/MM/YYYY format

	Returns a tuple of the form:
		day, month, year

	If DD = 0 and MM = 0, returns
		None, None, year

	If DD = 0, returns:
		None, month, year

	If raw_date is a datetime object, retrieves the day, month, year from that object

	If raw_date is None or empty, returns:
		None, None, None

	All other cases raise a ValueError
	"""
	if raw_date == None:
		return None, None, None

	if type(raw_date) == datetime:
		return raw_date.day, raw_date.month, raw_date.year

	if type(raw_date) in (str, text_type):
		parts = raw_date.split('/')
		if len(parts) != 3:
			raise ValueError("Invalid date format (expected DD/MM/YYYY)")

		try:
			d, m, y = [int(item) for item in parts]
		except ValueError:
			raise ValueError("Invalid date format (expected DD/MM/YYYY)")

		if y < 1900:
			log.info("Year %s earlier than 1900 (please ensure this is not a typo)" % y)
		if y > date.today().year:
			raise ValueError("Invalid year: %s" % y)
		if m == 0 and d == 0:
			return None, None, y
		if m < 1 or m > 12:
			raise ValueError("Invalid month: %s" % m)
		if d == 0:
			return None, m, y

		# Will raise an error if date is not valid
		date(y, m, d)

		return d, m, y

	raise ValueError("Unable to process date: %s" % raw_date)

def parse_time(raw_time):
	"""
	Parses a time in HH:MM:SS format

	Returns a time object

	If raw_time is already a time object, it is returned as-is

	If raw_time is None or empty, returns None
	"""
	if raw_time == None:
		return None

	if type(raw_time) == time:
		return raw_time

	if type(raw_time) in (str, text_type):
		return datetime.strptime(raw_time, '%H:%M:%S').time()

	raise ValueError("Unable to process time: %s" % raw_time)

def normalize(s):
	"""
	Normalizes a string:

	 - replaces any sequence of whitespace characters (including non-breaking space) with a single space character
	"""
	if s == None:
		return None
	else:
		return re.sub(r'\s+', ' ', s)

def quoted_strings(s):
	return ", ".join(["'%s'" % x for x in s])

if __name__ == '__main__':
	main()
